"""Vista: Reportes — ventas + inventario + pérdidas, exportable"""
import flet as ft
from datetime import datetime, timedelta
from core import ventas as vtas, productos as prods, perdidas as perd
from theme import Colors
from utils.dialogs import msg
import os
from openpyxl import Workbook


class ReportesView:
    def __init__(self, page):
        self.page = page
        self._filtradas = []

    def build(self):
        r_hoy = vtas.resumen_hoy()
        r_mes = vtas.resumen([v for v in vtas.listar() if vtas._fecha_valida(v, datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0))])

        self.stats_ganancias = ft.Text(
            f"💰 Ganancias:  Hoy: ₡{r_hoy['ganancia']:,.0f}  |  Este mes: ₡{r_mes['ganancia']:,.0f}",
            size=16, weight=ft.FontWeight.BOLD, color=Colors.SUCCESS
        )
        self._export_btn = ft.IconButton(ft.Icons.DOWNLOAD, icon_size=22,
            tooltip="Exportar a Excel", on_click=lambda e: self._exportar(),
            style=ft.ButtonStyle(color=ft.Colors.GREEN_700))
        self.stats = ft.Text("", size=14, weight=ft.FontWeight.W_600, color=Colors.PRIMARY)
        self.stats_inv = ft.Text("", size=13, color=Colors.TEXT_SEC)

        dd = ft.Dropdown(
            options=[ft.dropdown.Option(p) for p in
                     ["Hoy", "Esta semana", "Este mes", "Últimos 3 meses", "Todo"]],
            value="Hoy", width=180,
            on_select=lambda e: self._filtrar(e.control.value),
        )

        inventario_tabla = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Producto", size=10, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Stock", size=10, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Compra", size=10, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Venta", size=10, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Ganancia/u", size=10, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Valor Total", size=10, weight=ft.FontWeight.BOLD)),
            ],
            rows=[],
            border_radius=8,
            horizontal_lines=ft.BorderSide(1, ft.Colors.GREY_300),
            column_spacing=12,
        )

        inv_rows = []
        total_valor = 0
        for p in prods.listar():
            ganancia_u = (p["precio_venta"] or 0) - (p["precio_compra"] or 0)
            valor_total = p["stock"] * p["precio_venta"] if p["stock"] else 0
            total_valor += valor_total
            inv_rows.append(ft.DataRow([
                ft.DataCell(ft.Text(p["nombre"])),
                ft.DataCell(ft.Text(str(p["stock"]),
                    color=Colors.DANGER if (p["stock"] or 0) <= 3 else Colors.TEXT_PRIM)),
                ft.DataCell(ft.Text(f"₡{p['precio_compra']:,.0f}")),
                ft.DataCell(ft.Text(f"₡{p['precio_venta']:,.0f}")),
                ft.DataCell(ft.Text(f"₡{ganancia_u:,.0f}", color=Colors.SUCCESS)),
                ft.DataCell(ft.Text(f"₡{valor_total:,.0f}")),
            ]))
        inventario_tabla.rows = inv_rows

        self.stats_inv.value = f"📊 Inventario: {len(prods.listar())} productos | Valor total: ₡{total_valor:,.0f}"

        self.tabla = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Fecha", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Hora", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Producto", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Cant", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Precio", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Pago", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Subtotal", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Ganancia", size=11, weight=ft.FontWeight.BOLD)),
            ],
            rows=[], border_radius=8,
            horizontal_lines=ft.BorderSide(1, ft.Colors.GREY_300),
            column_spacing=15,
        )

        self._filtrar("Hoy")
        return ft.Column([
            ft.Row([
                ft.Text("📈 Reportes", size=20, weight=ft.FontWeight.BOLD, color=Colors.PRIMARY),
                ft.Container(expand=True),
                self.stats_ganancias,
                dd,
                self._export_btn,
            ]),
            # Sección inventario
            ft.Text("Inventario Actual", size=16, weight=ft.FontWeight.W_600, color=Colors.PRIMARY),
            self.stats_inv,
            ft.Container(content=ft.Column([inventario_tabla], scroll=ft.ScrollMode.AUTO),
                height=200, bgcolor=ft.Colors.WHITE, border_radius=12, padding=10,
                shadow=ft.BoxShadow(spread_radius=1, blur_radius=8, color=Colors.CARD_SHADOW)),
            ft.Divider(height=2, color=Colors.ACCENT),
            # Sección ventas
            ft.Text("Ventas", size=16, weight=ft.FontWeight.W_600, color=Colors.PRIMARY),
            self.stats,
            ft.Container(content=ft.Column([self.tabla], scroll=ft.ScrollMode.AUTO),
                expand=True, bgcolor=ft.Colors.WHITE, border_radius=12, padding=10,
                shadow=ft.BoxShadow(spread_radius=1, blur_radius=8, color=Colors.CARD_SHADOW)),
        ], expand=True)

    def _filtrar(self, periodo):
        ahora = datetime.now()
        hoy = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
        if periodo == "Hoy":
            inicio = hoy
        elif periodo == "Esta semana":
            inicio = hoy - timedelta(days=hoy.weekday())
        elif periodo == "Este mes":
            inicio = hoy.replace(day=1)
        elif periodo == "Últimos 3 meses":
            inicio = hoy - timedelta(days=90)
        else:
            inicio = datetime(2000, 1, 1)

        self._filtradas = [v for v in vtas.listar() if vtas._fecha_valida(v, inicio)]
        self.tabla.rows.clear()
        r = vtas.resumen(self._filtradas)

        self.stats.value = (
            f"📊 {periodo}: {len(self._filtradas)} ventas | "
            f"₡{r['subtotal']:,.0f} | 💰 ₡{r['ganancia']:,.0f} gana\n"
            f"💵 Efectivo: ₡{r['efectivo']:,.0f}  |  💳 SINPE: ₡{r['sinpe']:,.0f}"
        )

        for v in reversed(self._filtradas):
            self.tabla.rows.append(ft.DataRow([
                ft.DataCell(ft.Text(v["fecha"] if v["fecha"] else "")),
                ft.DataCell(ft.Text(v.get("hora", ""))),
                ft.DataCell(ft.Text(v["producto_nombre"])),
                ft.DataCell(ft.Text(str(v["cantidad"]))),
                ft.DataCell(ft.Text(f"₡{v['precio_unitario']:,.0f}")),
                ft.DataCell(ft.Text(v.get("metodo_pago", ""))),
                ft.DataCell(ft.Text(f"₡{v['subtotal']:,.0f}")),
                ft.DataCell(ft.Text(f"₡{v['ganancia']:,.0f}", color=Colors.SUCCESS)),
            ]))
        self.page.update()

    def _exportar(self):
        if not self._filtradas:
            msg(self.page, "Sin datos", "No hay ventas en este período", "error")
            return

        def on_result(e: ft.FilePickerResultEvent):
            if not e.path: return
            wb = Workbook()

            # Hoja 1: Ventas
            ws = wb.active; ws.title = "Ventas"
            ws.append(["Fecha","Hora","Producto","Cantidad","Precio","Método","Subtotal","Ganancia"])
            for v in self._filtradas:
                ws.append([v["fecha"], v.get("hora",""), v["producto_nombre"],
                    v["cantidad"], v["precio_unitario"], v.get("metodo_pago",""),
                    v["subtotal"], v["ganancia"]])
            tf = len(self._filtradas) + 2
            ws.cell(row=tf, column=6, value="TOTALES")
            ws.cell(row=tf, column=7, value=sum(v["subtotal"] or 0 for v in self._filtradas))
            ws.cell(row=tf, column=8, value=sum(v["ganancia"] or 0 for v in self._filtradas))

            # Hoja 2: Inventario
            ws2 = wb.create_sheet("Inventario")
            ws2.append(["Producto","Stock","Precio Compra","Precio Venta","Ganancia/u","Valor Total"])
            for p in prods.listar():
                gan_u = (p["precio_venta"] or 0) - (p["precio_compra"] or 0)
                val_t = (p["stock"] or 0) * (p["precio_venta"] or 0)
                ws2.append([p["nombre"], p["stock"], p["precio_compra"],
                          p["precio_venta"], gan_u, val_t])

            for ws_i in [ws, ws2]:
                for col in ws_i.columns:
                    ws_i.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 3
            wb.save(e.path)
            self.page.show_dialog(ft.AlertDialog(
                title=ft.Text("✅ Exportado"),
                content=ft.Text(f"Guardado: {os.path.basename(e.path)}"),
                actions=[ft.TextButton("OK", on_click=lambda e: self.page.pop_dialog())]
            ))
            self.page.update()

        picker = ft.FilePicker(on_result=on_result)
        self.page.overlay.append(picker)
        self.page.update()
        picker.save_file(file_name=f"reporte_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                         allowed_extensions=["xlsx"])
