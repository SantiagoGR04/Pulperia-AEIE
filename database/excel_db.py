"""Capa de datos — organización por año/mes + migración automática"""
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from config import DATA_DIR

MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
         "Julio","Agosto","Setiembre","Octubre","Noviembre","Diciembre"]

_COLS_PROD = ["id","nombre","precio_compra","precio_venta","stock","categoria"]
_COLS_PERD = ["id","producto_id","producto_nombre","cantidad","monto_estimado","motivo","fecha"]
_COLS_VENT = ["id","producto_id","producto_nombre","cantidad","precio_unitario","subtotal","ganancia","metodo_pago","fecha","hora"]

_PROD_FILE = os.path.join(DATA_DIR, "productos.xlsx")

# ─── helpers ────────────────────────────────────────────
def _año_mes():
    a = datetime.now()
    return a.year, a.month

def _mes_str(n):
    return MESES[n - 1]

def _parse_fecha(s):
    if not s:
        return _año_mes()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f"):
        try:
            d = datetime.strptime(s, fmt)
            return d.year, d.month
        except:
            continue
    return _año_mes()

def _yrf(tipo, año=None):
    return os.path.join(DATA_DIR, f"{tipo}_{año or _año_mes()[0]}.xlsx")

def _sheet(wb, name, cols):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(title=name)
    for i, c in enumerate(cols, 1):
        ws.cell(row=1, column=i, value=c)
    ws.freeze_panes = "A2"
    return ws

def _abrir(path, sheet_nombre, cols):
    os.makedirs(DATA_DIR, exist_ok=True)
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        for sn in list(wb.sheetnames):
            del wb[sn]
    ws = _sheet(wb, sheet_nombre, cols)
    return wb, ws

def _leer(path, cols):
    if not os.path.exists(path):
        return []
    try:
        wb = load_workbook(path)
    except Exception:
        return []
    rows = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            d = {}
            for i, c in enumerate(cols, 1):
                d[c] = ws.cell(row=r, column=i).value
            if d[cols[0]] is not None:
                rows.append(d)
    wb.close()
    return rows

def _leer_todos(prefix, cols):
    res = []
    if not os.path.isdir(DATA_DIR):
        return res
    for f in sorted(os.listdir(DATA_DIR)):
        if f.startswith(prefix) and f.endswith(".xlsx"):
            res.extend(_leer(os.path.join(DATA_DIR, f), cols))
    return res

# ─── productos (vivo, archivo único) ────────────────────
def _prod_wb():
    os.makedirs(DATA_DIR, exist_ok=True)
    if os.path.exists(_PROD_FILE):
        wb = load_workbook(_PROD_FILE)
    else:
        wb = Workbook()
        for sn in list(wb.sheetnames):
            del wb[sn]
    ws = _sheet(wb, "productos", _COLS_PROD)
    return wb, ws

def insertar_producto(d):
    wb, ws = _prod_wb()
    ws.append(d)
    wb.save(_PROD_FILE)

def actualizar_producto(pid, col, val):
    wb, ws = _prod_wb()
    ci = _COLS_PROD.index(col) + 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == pid:
            ws.cell(row=r, column=ci, value=val)
            break
    wb.save(_PROD_FILE)

def eliminar_producto(pid):
    wb, ws = _prod_wb()
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == pid:
            ws.delete_rows(r)
            break
    wb.save(_PROD_FILE)

def todos_productos():
    return _leer(_PROD_FILE, _COLS_PROD)

# ─── pérdidas (por año/mes) ─────────────────────────────
def insertar_perdida(d):
    año, mes = _parse_fecha(d[6])
    path = _yrf("perdidas", año)
    wb, ws = _abrir(path, _mes_str(mes), _COLS_PERD)
    ws.append(d)
    wb.save(path)

def todos_perdidas(año=None):
    if año:
        return _leer(_yrf("perdidas", año), _COLS_PERD)
    return _leer_todos("perdidas_", _COLS_PERD)

# ─── ventas (por año/mes) ───────────────────────────────
def insertar_venta(d):
    año, mes = _parse_fecha(d[8])
    path = _yrf("ventas", año)
    wb, ws = _abrir(path, _mes_str(mes), _COLS_VENT)
    ws.append(d)
    wb.save(path)

def todas_ventas(año=None):
    if año:
        return _leer(_yrf("ventas", año), _COLS_VENT)
    return _leer_todos("ventas_", _COLS_VENT)

# ─── migración (inventario.xlsx / ventas.xlsx → nuevo) ──
def _migrar():
    viejo_inv = os.path.join(DATA_DIR, "inventario.xlsx")
    if os.path.exists(viejo_inv):
        try:
            wb = load_workbook(viejo_inv)
            for sn in wb.sheetnames:
                ws = wb[sn]
                if sn == "productos":
                    for r in range(2, ws.max_row + 1):
                        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                        if row[0] is not None:
                            insertar_producto(row[:len(_COLS_PROD)])
                elif sn == "perdidas":
                    for r in range(2, ws.max_row + 1):
                        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                        if row[0] is not None:
                            insertar_perdida(row[:len(_COLS_PERD)])
            wb.close()
            os.rename(viejo_inv, viejo_inv + ".bak")
        except Exception as e:
            print(f"[Migración] inventario.xlsx → {e}")

    viejo_ven = os.path.join(DATA_DIR, "ventas.xlsx")
    if os.path.exists(viejo_ven):
        try:
            wb = load_workbook(viejo_ven)
            if "ventas" in wb.sheetnames:
                ws = wb["ventas"]
                for r in range(2, ws.max_row + 1):
                    row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                    if row[0] is not None:
                        insertar_venta(row[:len(_COLS_VENT)])
            wb.close()
            os.rename(viejo_ven, viejo_ven + ".bak")
        except Exception as e:
            print(f"[Migración] ventas.xlsx → {e}")

# ─── init ───────────────────────────────────────────────
def inicializar():
    os.makedirs(DATA_DIR, exist_ok=True)
    _migrar()
