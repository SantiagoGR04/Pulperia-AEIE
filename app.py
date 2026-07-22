"""Pulpería IE — App de escritorio moderna con Flet
Temática: Escuela de Ingeniería Eléctrica (azules, amarillos/grises técnicos)
"""
import flet as ft
from database.excel_db import inicializar, todos, insertar, actualizar, eliminar, obtener_por_id
from config import APP_NAME, APP_WIDTH, APP_HEIGHT, CATEGORIAS
from datetime import datetime, timedelta
from tkinter import filedialog, Tk
from openpyxl import Workbook
import threading
import os

# ─── PALETA IE ───────────────────────────────────────
class Colors:
    PRIMARY    = "#0D47A1"   # Azul profundo
    PRIMARY_DK = "#002171"   # Azul más oscuro
    PRIMARY_LT = "#5472D3"   # Azul claro
    ACCENT     = "#FFC107"   # Ámbar / dorado
    ACCENT_DK  = "#FF8F00"   # Ámbar oscuro
    SURFACE    = "#FFFFFF"
    BG         = "#F0F2F5"
    TEXT_PRIM  = "#1A1A2E"
    TEXT_SEC   = "#546E7A"
    SUCCESS    = "#2E7D32"
    DANGER     = "#C62828"
    WARNING    = "#E65100"
    CARD_SHADOW = "#1A000000"


def _build_icon(name, color=None, size=20):
    return ft.Icon(name=name, color=color or Colors.ACCENT, size=size)


# ─── DIALOGOS REUTILIZABLES ──────────────────────────

def _msg(page, titulo, texto, tipo="info"):
    c = ft.Colors.GREEN if tipo == "ok" else ft.Colors.RED if tipo == "error" else ft.Colors.BLUE
    d = ft.AlertDialog(
        title=ft.Text(titulo, color=c),
        content=ft.Text(texto),
        actions=[ft.TextButton("OK", on_click=lambda e: page.close(d))],
    )
    page.open(d)


# ─── VISTA: PRODUCTOS ────────────────────────────────

class ProductosView:
    def __init__(self, page):
        self.page = page
        self.search_text = ""

    def build(self):
        self.tabla = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("ID", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Producto", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Compra (₡)", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Venta (₡)", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Stock", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Categoría", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Acciones", size=11, weight=ft.FontWeight.BOLD)),
            ],
            rows=[],
            border_radius=8,
            horizontal_lines=ft.BorderSide(1, ft.Colors.GREY_300),
            column_spacing=25,
            width=1000,
        )

        self.search_input = ft.TextField(
            hint_text="Buscar producto…",
            prefix_icon=ft.Icons.SEARCH,
            width=300,
            height=40,
            on_change=lambda e: self._cargar(),
            border_radius=20,
            bgcolor=ft.Colors.WHITE,
        )

        self._cargar()
        return ft.Column([
            ft.Row([
                ft.Text("📦 Gestión de Productos", size=20, weight=ft.FontWeight.BOLD, color=Colors.PRIMARY),
                ft.Container(expand=True),
                self.search_input,
                ft.FloatingActionButton(
                    icon=ft.Icons.ADD,
                    text="Nuevo",
                    bgcolor=Colors.PRIMARY,
                    color=ft.Colors.WHITE,
                    on_click=lambda e: self._formulario(),
                ),
            ], alignment=ft.MainAxisAlignment.START),
            ft.Divider(height=2, color=Colors.ACCENT),
            ft.Container(
                content=ft.Column([self.tabla], scroll=ft.ScrollMode.AUTO),
                expand=True,
                bgcolor=ft.Colors.WHITE,
                border_radius=12,
                padding=10,
                shadow=ft.BoxShadow(spread_radius=1, blur_radius=8, color=Colors.CARD_SHADOW),
            ),
        ], expand=True)

    def _cargar(self):
        self.tabla.rows.clear()
        filtro = self.search_input.value.strip().lower() if self.search_input.value else ""
        prods = todos("productos")
        for p in prods:
            if filtro and filtro not in p["nombre"].lower():
                continue
            pid = p["id"]
            self.tabla.rows.append(ft.DataRow([
                ft.DataCell(ft.Text(str(pid))),
                ft.DataCell(ft.Text(p["nombre"], weight=ft.FontWeight.W_500)),
                ft.DataCell(ft.Text(f"₡{p['precio_compra']:,.0f}")),
                ft.DataCell(ft.Text(f"₡{p['precio_venta']:,.0f}")),
                ft.DataCell(ft.Text(str(p["stock"]),
                    color=Colors.DANGER if (p["stock"] or 0) == 0 else Colors.TEXT_PRIM)),
                ft.DataCell(ft.Text(p["categoria"] or "")),
                ft.DataCell(ft.Row([
                    ft.IconButton(ft.Icons.EDIT, icon_size=18, data=pid,
                        on_click=lambda e: self._formulario(int(e.control.data))),
                    ft.IconButton(ft.Icons.DELETE_OUTLINE, icon_size=18, data=pid,
                        on_click=lambda e: self._eliminar(int(e.control.data)),
                        icon_color=Colors.DANGER),
                ])),
            ]))
        self.page.update()

    def _eliminar(self, pid):
        d = ft.AlertDialog(
            title=ft.Text("Confirmar"),
            content=ft.Text("¿Eliminar este producto?"),
            actions=[
                ft.TextButton("Cancelar", on_click=lambda e: self.page.close(d)),
                ft.TextButton("Eliminar", on_click=lambda e: (
                    eliminar("productos", pid), self.page.close(d), self._cargar(), self.page.update()
                ), style=ft.ButtonStyle(color=Colors.DANGER)),
            ],
        )
        self.page.open(d)

    def _formulario(self, pid=None):
        prod = obtener_por_id("productos", pid) if pid else None
        editando = prod is not None

        e_nombre = ft.TextField(label="Nombre", value=prod["nombre"] if editando else "", width=300)
        e_compra = ft.TextField(label="Precio compra (₡)", value=str(prod["precio_compra"]) if editando else "", width=300, keyboard_type=ft.KeyboardType.NUMBER)
        e_venta  = ft.TextField(label="Precio venta (₡)", value=str(prod["precio_venta"]) if editando else "", width=300, keyboard_type=ft.KeyboardType.NUMBER)
        e_stock  = ft.TextField(label="Stock", value=str(prod["stock"]) if editando else "", width=300, keyboard_type=ft.KeyboardType.NUMBER)
        e_cat    = ft.Dropdown(
            label="Categoría",
            options=[ft.dropdown.Option(c) for c in CATEGORIAS],
            value=prod["categoria"] if editando else CATEGORIAS[0],
            width=300,
        )

        def guardar(e):
            try:
                nom = e_nombre.value.strip()
                if not nom:
                    _msg(self.page, "Error", "Nombre requerido", "error")
                    return
                pc = float(e_compra.value)
                pv = float(e_venta.value)
                st = int(e_stock.value)
                if editando:
                    actualizar("productos", pid, "nombre", nom)
                    actualizar("productos", pid, "precio_compra", pc)
                    actualizar("productos", pid, "precio_venta", pv)
                    actualizar("productos", pid, "stock", st)
                    actualizar("productos", pid, "categoria", e_cat.value)
                else:
                    prods = todos("productos")
                    nid = max((p["id"] for p in prods), default=0) + 1
                    insertar("productos", [nid, nom, pc, pv, st, e_cat.value])
                self.page.close(dlg)
                self._cargar()
                self.page.update()
            except ValueError:
                _msg(self.page, "Error", "Precios y stock deben ser números", "error")

        dlg = ft.AlertDialog(
            title=ft.Text("Editar Producto" if editando else "Nuevo Producto"),
            content=ft.Column([
                e_nombre, e_compra, e_venta, e_stock, e_cat,
                ft.ElevatedButton("💾 Guardar", on_click=guardar, bgcolor=Colors.PRIMARY, color=ft.Colors.WHITE),
            ], tight=True, scroll=ft.ScrollMode.AUTO),
            actions=[ft.TextButton("Cancelar", on_click=lambda e: self.page.close(dlg))],
        )
        self.page.open(dlg)
        dlg.update()


# ─── VISTA: PÉRDIDAS ────────────────────────────────

class PerdidasView:
    def __init__(self, page):
        self.page = page

    def build(self):
        self.tabla = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Producto", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Cant", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Monto", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Motivo", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Fecha", size=11, weight=ft.FontWeight.BOLD)),
            ],
            rows=[],
            border_radius=8,
            horizontal_lines=ft.BorderSide(1, ft.Colors.GREY_300),
            column_spacing=25,
            width=1000,
        )

        self._cargar_tabla()

        return ft.Column([
            ft.Row([
                ft.Text("⚠️ Registro de Pérdidas", size=20, weight=ft.FontWeight.BOLD, color=Colors.DANGER),
                ft.Container(expand=True),
                ft.FloatingActionButton(
                    icon=ft.Icons.ADD,
                    text="Nueva Pérdida",
                    bgcolor=Colors.DANGER,
                    color=ft.Colors.WHITE,
                    on_click=lambda e: self._formulario(),
                ),
            ]),
            ft.Divider(height=2, color=Colors.ACCENT),
            ft.Container(
                content=ft.Column([self.tabla], scroll=ft.ScrollMode.AUTO),
                expand=True,
                bgcolor=ft.Colors.WHITE,
                border_radius=12,
                padding=10,
                shadow=ft.BoxShadow(spread_radius=1, blur_radius=8, color=Colors.CARD_SHADOW),
            ),
        ], expand=True)

    def _cargar_tabla(self):
        self.tabla.rows.clear()
        perdidas = todos("perdidas")
        for p in reversed(perdidas[-50:]):
            fecha = p["fecha"][:10] if p["fecha"] else ""
            self.tabla.rows.append(ft.DataRow([
                ft.DataCell(ft.Text(p["producto_nombre"] or "")),
                ft.DataCell(ft.Text(str(p["cantidad"] or ""))),
                ft.DataCell(ft.Text(f"₡{p['monto_estimado']:,.0f}" if p['monto_estimado'] else "₡0")),
                ft.DataCell(ft.Text(p["motivo"] or "")),
                ft.DataCell(ft.Text(fecha)),
            ]))
        self.page.update()

    def _formulario(self):
        prods = todos("productos")
        opciones = [f"{p['id']} - {p['nombre']}" for p in prods]
        opciones.append("Otro (nombre libre)")

        e_prod = ft.Dropdown(
            label="Producto",
            options=[ft.dropdown.Option(o) for o in opciones],
            width=350,
        )
        e_cant = ft.TextField(label="Cantidad estimada", width=350, keyboard_type=ft.KeyboardType.NUMBER)
        e_monto = ft.TextField(label="Monto estimado (₡)", width=350, keyboard_type=ft.KeyboardType.NUMBER)
        e_motivo = ft.Dropdown(
            label="Motivo",
            options=[ft.dropdown.Option(m) for m in [
                "No se registró venta", "Robo/Hurto", "Daño/Caducó",
                "Error de inventario", "Otro"
            ]],
            value="No se registró venta",
            width=350,
        )

        def guardar(e):
            txt = e_prod.value or ""
            pid = None
            pnom = txt
            try:
                pid = int(txt.split(" - ")[0])
                prod = obtener_por_id("productos", pid)
                if prod:
                    pnom = prod["nombre"]
            except:
                pnom = txt.replace("Otro (nombre libre)", "Producto no identificado") if "Otro" in txt else txt

            try:
                cant = int(e_cant.value) if e_cant.value else 0
            except:
                cant = 0
            try:
                monto = float(e_monto.value) if e_monto.value else 0
            except:
                monto = 0

            perdidas = todos("perdidas")
            nid = max((p["id"] for p in perdidas), default=0) + 1
            insertar("perdidas", [nid, pid, pnom, cant, monto, e_motivo.value, datetime.now().isoformat()])
            self.page.close(dlg)
            self._cargar_tabla()
            dlg.update()

        dlg = ft.AlertDialog(
            title=ft.Text("Registrar Pérdida"),
            content=ft.Column([
                e_prod, e_cant, e_monto, e_motivo,
                ft.ElevatedButton("💾 Guardar", on_click=guardar, bgcolor=Colors.DANGER, color=ft.Colors.WHITE),
            ], tight=True, scroll=ft.ScrollMode.AUTO),
            actions=[ft.TextButton("Cancelar", on_click=lambda e: self.page.close(dlg))],
        )
        self.page.open(dlg)


# ─── VISTA: REPORTES ────────────────────────────────

class ReportesView:
    def __init__(self, page):
        self.page = page
        self.periodo = "Hoy"
        self._filtradas = []

    def build(self):
        self.tabla = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Fecha", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Producto", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Cant", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Precio", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Subtotal", size=11, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("Ganancia", size=11, weight=ft.FontWeight.BOLD)),
            ],
            rows=[],
            border_radius=8,
            horizontal_lines=ft.BorderSide(1, ft.Colors.GREY_300),
            column_spacing=20,
            width=1000,
        )

        self.stats_text = ft.Text("", size=14, weight=ft.FontWeight.W_600, color=Colors.PRIMARY)

        dd_periodo = ft.Dropdown(
            options=[ft.dropdown.Option(p) for p in
                     ["Hoy", "Esta semana", "Este mes", "Últimos 3 meses", "Todo"]],
            value="Hoy",
            width=180,
            on_change=lambda e: self._filtrar(e.control.value),
        )

        self._filtrar("Hoy")
        return ft.Column([
            ft.Row([
                ft.Text("📈 Reportes", size=20, weight=ft.FontWeight.BOLD, color=Colors.PRIMARY),
                ft.Container(expand=True),
                dd_periodo,
                ft.ElevatedButton("💾 Exportar Excel", icon=ft.Icons.FILE_DOWNLOAD,
                    bgcolor=Colors.SUCCESS, color=ft.Colors.WHITE,
                    on_click=lambda e: self._exportar()),
            ]),
            self.stats_text,
            ft.Divider(height=2, color=Colors.ACCENT),
            ft.Container(
                content=ft.Column([self.tabla], scroll=ft.ScrollMode.AUTO),
                expand=True,
                bgcolor=ft.Colors.WHITE,
                border_radius=12,
                padding=10,
                shadow=ft.BoxShadow(spread_radius=1, blur_radius=8, color=Colors.CARD_SHADOW),
            ),
        ], expand=True)

    def _filtrar(self, periodo):
        self.periodo = periodo
        ahora = datetime.now()
        hoy = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
        if periodo == "Hoy":
            inicio = hoy
        elif periodo == "Esta semana":
            inicio = hoy - timedelta(days=hoy.weekday())
        elif periodo == "Este mes":
            inicio = hoy.replace(day=1)
        elif periodo == "Últimos 3 meses":
            inicio = hoy - timedelta(days=90)
        else:
            inicio = datetime(2000, 1, 1)

        ventas = todos("ventas")
        self._filtradas = []
        for v in ventas:
            try:
                fv = datetime.fromisoformat(v["fecha"])
                if fv >= inicio:
                    self._filtradas.append(v)
            except:
                continue

        self.tabla.rows.clear()
        total_sub = sum(v["subtotal"] or 0 for v in self._filtradas)
        total_gan = sum(v["ganancia"] or 0 for v in self._filtradas)
        total_cant = sum(v["cantidad"] or 0 for v in self._filtradas)

        self.stats_text.value = f"📊 {periodo}: {len(self._filtradas)} ventas | {total_cant} unidades | ₡{total_sub:,.0f} en ventas | 💰 Ganancia: ₡{total_gan:,.0f}"

        for v in reversed(self._filtradas):
            fecha = v["fecha"][:10] if v["fecha"] else ""
            self.tabla.rows.append(ft.DataRow([
                ft.DataCell(ft.Text(fecha)),
                ft.DataCell(ft.Text(v["producto_nombre"])),
                ft.DataCell(ft.Text(str(v["cantidad"]))),
                ft.DataCell(ft.Text(f"₡{v['precio_unitario']:,.0f}")),
                ft.DataCell(ft.Text(f"₡{v['subtotal']:,.0f}")),
                ft.DataCell(ft.Text(f"₡{v['ganancia']:,.0f}", color=Colors.SUCCESS)),
            ]))
        self.page.update()

    def _exportar(self):
        if not self._filtradas:
            _msg(self.page, "Sin datos", "No hay ventas en este período", "error")
            return

        # Usar threading para no bloquear Flet con el diálogo de archivos
        def exportar_thread():
            root = Tk()
            root.withdraw()
            filename = filedialog.asksaveasfilename(
                title="Guardar reporte",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"reporte_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            )
            root.destroy()
            if not filename:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte"
            ws.append(["Fecha", "Producto", "Cantidad", "Precio Unitario", "Subtotal", "Ganancia"])
            for v in self._filtradas:
                ws.append([
                    v["fecha"][:10] if v["fecha"] else "",
                    v["producto_nombre"], v["cantidad"],
                    v["precio_unitario"], v["subtotal"], v["ganancia"]
                ])
            tf = len(self._filtradas) + 2
            ws.cell(row=tf, column=4, value="TOTALES")
            ws.cell(row=tf, column=5, value=sum(v["subtotal"] or 0 for v in self._filtradas))
            ws.cell(row=tf, column=6, value=sum(v["ganancia"] or 0 for v in self._filtradas))
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 3
            wb.save(filename)
            self.page.open(ft.SnackBar(ft.Text(f"✅ Reporte guardado: {os.path.basename(filename)}"),
                                        bgcolor=Colors.SUCCESS))

        threading.Thread(target=exportar_thread, daemon=True).start()


# ─── APP PRINCIPAL ───────────────────────────────────

def main(page: ft.Page):
    page.title = APP_NAME
    page.theme_mode = ft.ThemeMode.LIGHT
    page.bgcolor = Colors.BG
    page.padding = 0
    page.window_width = APP_WIDTH
    page.window_height = APP_HEIGHT
    page.window_min_width = 900
    page.window_min_height = 600
    page.window_top = 100
    page.window_left = 100

    page.theme = ft.Theme(
        color_scheme=ft.ColorScheme(
            primary=Colors.PRIMARY,
            secondary=Colors.ACCENT,
            surface=Colors.SURFACE,
        ),
        font_family="Segoe UI",
    )

    inicializar()

    content_area = ft.Container(expand=True, padding=20)

    def cambiar_vista(nombre):
        content_area.content = None
        content_area.update()
        if nombre == "dashboard":
            content_area.content = DashboardView(page).build()
        elif nombre == "productos":
            content_area.content = ProductosView(page).build()
        elif nombre == "perdidas":
            content_area.content = PerdidasView(page).build()
        elif nombre == "reportes":
            content_area.content = ReportesView(page).build()
        content_area.update()

    # Sidebar
    sidebar = ft.Container(
        width=220,
        bgcolor=Colors.PRIMARY_DK,
        padding=0,
        content=ft.Column([
            ft.Container(
                content=ft.Column([
                    ft.Icon(ft.Icons.BOLT, size=36, color=Colors.ACCENT),
                    ft.Text("Pulpería IE", size=20, weight=ft.FontWeight.BOLD, color=ft.Colors.WHITE),
                    ft.Text("Ing. Eléctrica", size=12, color=ft.Colors.GREY_400),
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                padding=20,
                bgcolor=Colors.PRIMARY,
            ),
            ft.Divider(height=1, color=Colors.ACCENT),
            ft.Container(expand=True, content=ft.Column([
                ft.ListTile(
                    leading=ft.Icon(ft.Icons.DASHBOARD, color=ft.Colors.WHITE),
                    title=ft.Text("Dashboard", color=ft.Colors.WHITE),
                    on_click=lambda e: cambiar_vista("dashboard"),
                ),
                ft.ListTile(
                    leading=ft.Icon(ft.Icons.INVENTORY_2, color=ft.Colors.WHITE70),
                    title=ft.Text("Productos", color=ft.Colors.WHITE),
                    on_click=lambda e: cambiar_vista("productos"),
                ),
                ft.ListTile(
                    leading=ft.Icon(ft.Icons.WARNING_AMBER, color=ft.Colors.WHITE70),
                    title=ft.Text("Pérdidas", color=ft.Colors.WHITE),
                    on_click=lambda e: cambiar_vista("perdidas"),
                ),
                ft.ListTile(
                    leading=ft.Icon(ft.Icons.BAR_CHART, color=ft.Colors.WHITE70),
                    title=ft.Text("Reportes", color=ft.Colors.WHITE),
                    on_click=lambda e: cambiar_vista("reportes"),
                ),
            ], spacing=0)),
            ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.ELECTRIC_BOLT, size=14, color=Colors.ACCENT),
                    ft.Text("IE-UCR", size=11, color=ft.Colors.GREY_500),
                ], alignment=ft.MainAxisAlignment.CENTER),
                padding=10,
            ),
        ]),
    )

    page.add(ft.Row([
        sidebar,
        content_area,
    ], expand=True, spacing=0))

    cambiar_vista("dashboard")
    page.navegar = cambiar_vista


# ─── VISTA: DASHBOARD ───────────────────────────────

class DashboardView:
    def __init__(self, page):
        self.page = page

    def _navegar_a(self, destino):
        self.page.navegar(destino)

    def build(self):
        prods = todos("productos")
        ventas = todos("ventas")
        perdidas = todos("perdidas")

        hoy = datetime.now().strftime("%Y-%m-%d")
        ventas_hoy = [v for v in ventas if v.get("fecha","").startswith(hoy)]
        ganancia_total = sum(v["ganancia"] or 0 for v in ventas)
        ventas_hoy_total = sum(v["subtotal"] or 0 for v in ventas_hoy)
        perdidas_total = sum(p["monto_estimado"] or 0 for p in perdidas)
        prod_bajo_stock = len([p for p in prods if (p["stock"] or 0) <= 3])

        def card(icon, label, valor, bg, icon_color=Colors.ACCENT):
            return ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Icon(icon, color=icon_color, size=28),
                        ft.Container(expand=True),
                    ]),
                    ft.Text(valor, size=28, weight=ft.FontWeight.BOLD, color=ft.Colors.WHITE),
                    ft.Text(label, size=13, color=ft.Colors.WHITE70),
                ], spacing=5),
                bgcolor=bg,
                padding=20,
                border_radius=16,
                expand=True,
                shadow=ft.BoxShadow(spread_radius=1, blur_radius=10, color=Colors.CARD_SHADOW),
            )

        return ft.Column([
            ft.Row([
                ft.Text("⚡ Dashboard", size=22, weight=ft.FontWeight.BOLD, color=Colors.PRIMARY),
                ft.Container(expand=True),
                ft.Text(datetime.now().strftime("%A, %d %B %Y"), size=13, color=Colors.TEXT_SEC),
            ]),
            ft.Divider(height=2, color=Colors.ACCENT),
            ft.Row([
                card(ft.Icons.INVENTORY, "Productos", str(len(prods)), Colors.PRIMARY),
                card(ft.Icons.TRENDING_UP, "Ventas hoy", f"₡{ventas_hoy_total:,.0f}", Colors.SUCCESS),
                card(ft.Icons.ACCOUNT_BALANCE, "Ganancia total", f"₡{ganancia_total:,.0f}", "#1565C0"),
                card(ft.Icons.WARNING, f"Pérdidas", f"₡{perdidas_total:,.0f}", Colors.DANGER, icon_color=ft.Colors.WHITE),
            ], spacing=15),
            ft.Row([
                ft.Container(
                    expand=True,
                    content=ft.Column([
                        ft.Text("📋 Resumen Rápido", size=16, weight=ft.FontWeight.BOLD, color=Colors.PRIMARY),
                        ft.Divider(),
                        ft.Text(f"🔴 Productos con stock bajo (≤3): {prod_bajo_stock}",
                                color=Colors.DANGER if prod_bajo_stock > 0 else Colors.TEXT_SEC),
                        ft.Text(f"📦 Total de ventas registradas: {len(ventas)}"),
                        ft.Text(f"⚠️ Pérdidas registradas: {len(perdidas)}"),
                    ]),
                    bgcolor=ft.Colors.WHITE,
                    border_radius=12,
                    padding=15,
                    shadow=ft.BoxShadow(spread_radius=1, blur_radius=8, color=Colors.CARD_SHADOW),
                ),
                ft.Container(
                    expand=True,
                    content=ft.Column([
                        ft.Text("🎯 Accesos Rápidos", size=16, weight=ft.FontWeight.BOLD, color=Colors.PRIMARY),
                        ft.Divider(),
                        ft.ElevatedButton("➕ Nuevo Producto", icon=ft.Icons.ADD,
                            on_click=lambda e: self._navegar_a("productos"), width=200),
                        ft.ElevatedButton("📊 Ver Reportes", icon=ft.Icons.BAR_CHART,
                            on_click=lambda e: self._navegar_a("reportes"), width=200),
                    ], spacing=10),
                    bgcolor=ft.Colors.WHITE,
                    border_radius=12,
                    padding=15,
                    shadow=ft.BoxShadow(spread_radius=1, blur_radius=8, color=Colors.CARD_SHADOW),
                ),
            ]),
        ], expand=True)


ft.run(main)
